import json
import re
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from sklearn.decomposition import TruncatedSVD
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.manifold import TSNE
from sklearn.preprocessing import normalize


ROOT = Path(__file__).resolve().parents[1]
CONTENT_PATH = ROOT / "data" / "content.json"
SCATTER_PATH = ROOT / "data" / "scatter.json"
TAXONOMY_COLORS_PATH = ROOT / "data" / "taxonomy_colors.json"
TAXONOMY_XLSX_PATH = ROOT / "vis_data" / "taxonomy-new(1).xlsx"


TEXT_FIELDS = [
    "title",
    "authors",
    "keywords",
    "Keywords",
    "abstract",
    "anchor_summary",
    "dataset",
    "interaction_detail",
]

CATEGORY_FIELDS = [
    "salon",
    "task",
    "modality",
    "domain",
    "model",
    "vis_encoding",
    "specific_view",
    "interaction",
    "evaluation",
    "special_tags",
]

DRAWING_NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
}

TAXONOMY_TITLE_TO_KEY = {
    "Data Modality": "modality",
    "Model": "model",
    "Visualization": "vis_encoding",
    "Interaction": "interaction",
    "Evaluation": "evaluation",
    "Cross-modal Task": "task",
    "Application": "domain",
}


def as_list(value):
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    value = str(value).strip()
    return [value] if value else []


def clean_text(value):
    value = str(value or "")
    value = re.sub(r"<[^>]+>", " ", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def build_document(entry):
    parts = []

    for field in TEXT_FIELDS:
        text = clean_text(entry.get(field))
        if text:
            parts.append(text)

    # Repeat structured labels so the embedding respects survey taxonomy, not
    # only abstract wording.
    for field in CATEGORY_FIELDS:
        values = as_list(entry.get(field))
        if values:
            parts.extend(values * 3)

    return " ".join(parts) or clean_text(entry.get("title")) or entry.get("id", "")


def normalize_coordinates(coords):
    coords = np.asarray(coords, dtype=float)
    coords -= coords.mean(axis=0)

    cov = np.cov(coords, rowvar=False)
    eigvals, eigvecs = np.linalg.eigh(cov)
    eigvals[eigvals < 1e-9] = 1e-9
    coords = coords @ eigvecs @ np.diag(1 / np.sqrt(eigvals))
    coords -= coords.mean(axis=0)

    scale = np.percentile(np.abs(coords), 98, axis=0)
    scale[scale == 0] = 1
    coords = coords / scale * 5

    return np.clip(coords, -6, 6)


def hex_color(value):
    value = str(value or "").strip().upper()
    if len(value) == 6:
        return "#" + value
    if len(value) == 8:
        return "#" + value[-6:]
    return ""


def top_header_ranges(ws):
    ranges = []
    for merged in ws.merged_cells.ranges:
        if merged.min_row == 1 and merged.max_row == 1:
            title = ws.cell(1, merged.min_col).value
            if title:
                ranges.append((merged.min_col, merged.max_col, str(title).strip()))

    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if not value:
            continue
        if any(start <= col <= end for start, end, _title in ranges):
            continue
        ranges.append((col, col, str(value).strip()))

    return sorted(ranges, key=lambda item: item[0])


def taxonomy_colors_from_excel():
    if not TAXONOMY_XLSX_PATH.exists():
        return {}

    workbook = load_workbook(TAXONOMY_XLSX_PATH, data_only=False)
    worksheet = workbook["Sheet1"]
    header_ranges = top_header_ranges(worksheet)

    column_to_key = {}
    for start_col, end_col, title in header_ranges:
        key = TAXONOMY_TITLE_TO_KEY.get(title)
        if not key:
            continue
        for col in range(start_col, end_col + 1):
            column_to_key[col] = key

    counts = defaultdict(Counter)
    with zipfile.ZipFile(TAXONOMY_XLSX_PATH) as archive:
        root = ET.fromstring(archive.read("xl/drawings/drawing1.xml"))
        for anchor in root.findall("xdr:twoCellAnchor", DRAWING_NS):
            col = int(anchor.findtext("xdr:from/xdr:col", namespaces=DRAWING_NS)) + 1
            key = column_to_key.get(col)
            if not key:
                continue

            color = ""
            for fill in anchor.findall(".//a:solidFill", DRAWING_NS):
                clr = fill.find("a:srgbClr", DRAWING_NS)
                if clr is not None:
                    color = hex_color(clr.attrib.get("val"))
                    if color:
                        break

            if color:
                counts[key][color] += 1

    colors = {}
    for key, counter in counts.items():
        colors[key] = counter.most_common(1)[0][0]
    return colors


def main():
    entries = json.loads(CONTENT_PATH.read_text(encoding="utf-8"))
    documents = [build_document(entry) for entry in entries]

    vectorizer = TfidfVectorizer(
        lowercase=True,
        stop_words="english",
        token_pattern=r"(?u)\b[\w][\w\-]+\b",
        min_df=1,
        max_df=0.85,
        ngram_range=(1, 2),
        max_features=3000,
    )
    matrix = vectorizer.fit_transform(documents)

    svd_components = min(50, matrix.shape[0] - 1, matrix.shape[1] - 1)
    if svd_components >= 2:
        dense = TruncatedSVD(n_components=svd_components, random_state=42).fit_transform(matrix)
    else:
        dense = matrix.toarray()

    dense = normalize(dense)

    perplexity = min(30, max(5, (len(entries) - 1) // 3))
    coords = TSNE(
        n_components=2,
        perplexity=perplexity,
        learning_rate="auto",
        init="pca",
        metric="cosine",
        random_state=42,
        max_iter=1500,
    ).fit_transform(dense)

    coords = normalize_coordinates(coords)

    scatter_entries = []
    for entry, coord in zip(entries, coords):
        item = {
            "id": entry.get("id", ""),
            "title": entry.get("title", ""),
            "x": round(float(coord[0]), 6),
            "y": round(float(coord[1]), 6),
        }
        for field in CATEGORY_FIELDS:
            item[field] = entry.get(field, [])
        item["year"] = entry.get("year", "")
        item["venue"] = entry.get("venue", "")
        item["authors"] = entry.get("authors", "")
        item["model"] = entry.get("model", [])
        scatter_entries.append(item)

    SCATTER_PATH.write_text(
        json.dumps(scatter_entries, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    taxonomy_colors = taxonomy_colors_from_excel()
    TAXONOMY_COLORS_PATH.write_text(
        json.dumps(taxonomy_colors, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"Wrote {len(scatter_entries)} entries to {SCATTER_PATH}")
    print(f"Wrote taxonomy colors to {TAXONOMY_COLORS_PATH}")


if __name__ == "__main__":
    main()
