import json
import re
from collections import Counter, defaultdict
from io import BytesIO
from pathlib import Path

import openpyxl
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[1]
XLSX_DIR = ROOT / "vis_data"
DATA_DIR = ROOT / "data"
BIBTEX_DIR = ROOT / "bibtex"
THUMBS100_DIR = ROOT / "thumbs100"
THUMBS200_DIR = ROOT / "thumbs200"


COL = {
    "title": 2,
    "screenshot": 3,
    "year": 4,
    "venue": 5,
    "rank": 6,
    "citations": 7,
    "bibtex": 8,
    "abstract": 9,
    "keywords": 10,
    "modal_count": 11,
    "modality": 12,
    "dataset": 13,
    "domain": 14,
    "salon": 15,
    "task": 16,
    "model": 17,
    "anchor_summary": 18,
    "improved_model": 19,
    "fine_tuning": 20,
    "vis_encoding": 21,
    "specific_view": 22,
    "interaction": 23,
    "interaction_detail": 24,
    "evaluation": 25,
    "url": 26,
}


CATEGORY_GROUPS = [
    ("modality", "Data Modality", "Data Modality"),
    ("salon", "SALON", "SALON"),
    ("task", "Downstream Task", "Downstream Task"),
    ("vis_encoding", "Visual Encoding", "Visual Encoding"),
    ("interaction", "Interaction Technique", "Interaction Technique"),
    ("evaluation", "Evaluation Method", "Evaluation Method"),
    ("special_tags", "Special Tags", "Special Tags"),
    ("model", "Base Model", "Base Model"),
    ("domain", "Application Domain", "Application Domain"),
]

CATEGORY_TRANSLATIONS = {
    "自然图像-文本": "Natural Image-Language",
    "科学图像-文本": "Scientific Image-Language",
    "图表图像-文本": "Chart Image-Language",
    "可视分析系统-文本": "Visual System-Language",
    "自然视频-文本": "Natural Video-Language",
    "数据视频-文本": "Data Video-Language",
    "自然视频-文本-语音": "Natural Video-Language-Audio",
    "数据视频-文本-语音": "Data Video-Language-Audio",
    "自然图像-文本-语音": "Natural Image-Language-Audio",
    "图表图像-文本-语音": "Chart Image-Language-Audio",
    "自然视频-文本-可视化": "Natural Video-Language-Visualization",
    "文档-文本-可视化": "Document-Language-Visualization",
    "自然视频-可视化": "Natural Video-Visualization",
    "自然图像": "Natural Image",
    "语音-文本": "Speech-Language",
    "文本": "Language",
    "通用领域": "General",
    "医学生物领域": "Medical/Bio",
    "交通领域": "Traffic",
    "地理环境领域": "Geo/Environment",
    "社交媒体领域": "Social Media",
    "教育学习领域": "Education/Learning",
    "体育运动领域": "Sports/Exercise",
    "历史文物领域": "Humanities/Cultural Heritage",
    "艺术创作领域": "Art Design/Creative",
    "电子商务领域": "E-commerce/Retail",
    "视觉语言-表示": "Vision-Language Representation",
    "视觉语言-转译": "Vision-Language Translation",
    "视觉语言-对齐": "Vision-Language Alignment",
    "视觉语言-融合": "Vision-Language Fusion",
    "视觉语言-协同学习": "Vision-Language Co-learning",
    "基础模型可视解释": "Model Interpretability",
    "跨模态生成任务-图生文": "Image-to-Text",
    "跨模态生成任务-文生图": "Text-to-Image",
    "跨模态理解与推理任务-注释": "Annotation",
    "跨模态关联与匹配任务-检索": "Retrieval",
    "跨模态关联与匹配任务-匹配": "Matching",
    "跨模态关联与匹配任务-定位": "Grounding",
    "跨模态理解与推理任务-问答": "Question Answering",
    "跨模态理解与推理任务-对话": "Grounded Dialogue",
    "多模态协同分析任务": "Multimodal Collaborative Analysis",
    "跨模态理解与推理任务-分类": "Classification",
    "CNN架构": "CNN",
    "RNN架构": "RNN",
    "LSTM架构": "LSTM",
    "Transformer架构": "Transformer",
    "BLIP系列": "BLIP",
    "GPT系列": "GPT Series",
    "Gemini系列": "Gemini Series",
    "Qwen系列": "Qwen Series",
    "Claude系列": "Claude Series",
    "BERT系列": "BERT Series",
    "DeepSeek系列": "DeepSeek Series",
    "Gemma系列": "Gemma Series",
    "Baichuan系列": "Baichuan Series",
    "领域模型": "Domain-specific Model",
    "符号编码": "Symbolic Encoding",
    "离散型编码": "Discrete Encoding",
    "分布型编码": "Distribution Encoding",
    "时序型编码": "Temporal Encoding",
    "结构型编码": "Structural Encoding",
    "层级型编码": "Hierarchical Encoding",
    "关联操作": "Linking",
    "筛选操作": "Filtering",
    "视点操作": "Viewpoint Control",
    "What-If操作": "What-If",
    "记录操作": "Recording/History",
    "定性案例分析": "Case Study",
    "人类用户体验": "User Experience",
    "定量指标评估": "Quantitative Experiment",
    "人类用户表现": "User Performance",
    "其他": "Other",
}

CATEGORY_ORDER = {
    "salon": [
        "Vision-Language Representation",
        "Vision-Language Translation",
        "Vision-Language Alignment",
        "Vision-Language Fusion",
        "Vision-Language Co-learning",
    ],
    "task": [
        "Model Interpretability",
        "Image-to-Text",
        "Text-to-Image",
        "Question Answering",
        "Grounded Dialogue",
        "Classification",
        "Annotation",
        "Retrieval",
        "Matching",
        "Grounding",
        "Multimodal Collaborative Analysis",
    ],
    "modality": [
        "Natural Image-Language",
        "Scientific Image-Language",
        "Chart Image-Language",
        "Visual System-Language",
        "Natural Video-Language",
        "Data Video-Language",
        "Natural Video-Language-Audio",
        "Data Video-Language-Audio",
        "Natural Image-Language-Audio",
        "Chart Image-Language-Audio",
        "Natural Video-Language-Visualization",
        "Document-Language-Visualization",
        "Natural Video-Visualization",
        "Natural Image",
        "Speech-Language",
        "Language",
        "Other",
    ],
    "domain": [
        "General",
        "Medical/Bio",
        "Traffic",
        "Geo/Environment",
        "Social Media",
        "Education/Learning",
        "Sports/Exercise",
        "Humanities/Cultural Heritage",
        "Art Design/Creative",
        "E-commerce/Retail",
        "Other",
    ],
    "model": [
        "CLIP",
        "LLaVA",
        "GPT Series",
        "BLIP",
        "Gemini Series",
        "Qwen Series",
        "Claude Series",
        "DeepSeek Series",
        "Stable Diffusion",
        "DALL-E",
        "SDXL",
        "ViT",
        "SAM",
        "DINO",
        "Transformer",
        "BERT Series",
        "CNN",
        "RNN",
        "LSTM",
        "Gemma Series",
        "Baichuan Series",
        "LaMDA",
        "Domain-specific Model",
        "Other",
    ],
    "vis_encoding": [
        "Discrete Encoding",
        "Temporal Encoding",
        "Structural Encoding",
        "Hierarchical Encoding",
        "Distribution Encoding",
        "Symbolic Encoding",
    ],
    "interaction": [
        "Linking",
        "Filtering",
        "What-If",
        "Viewpoint Control",
        "Recording/History",
    ],
    "evaluation": [
        "Case Study",
        "User Performance",
        "User Experience",
        "Quantitative Experiment",
    ],
    "special_tags": ["Agent", "RAG"],
}


SPLIT_RE = re.compile(r"[,;；，、\n]+")
DOI_RE = re.compile(r"\b10\.\d{4,9}/[-._;()/:A-Z0-9]+", re.I)


def clean_text(value):
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"\s+\n", "\n", text)
    text = re.sub(r"\n\s+", "\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    return text.strip()


def split_values(value):
    text = clean_text(value)
    if not text or text == "无":
        return []
    values = []
    for item in SPLIT_RE.split(text):
        item = item.strip()
        if item and item != "无":
            values.append(item)
    return values


def translate_category_value(value):
    value = clean_text(value)
    return CATEGORY_TRANSLATIONS.get(value, value)


def translated_values(value):
    return [translate_category_value(item) for item in split_values(value)]


def slugify(value, prefix):
    text = clean_text(value).lower()
    text = re.sub(r"\s+", "-", text)
    text = re.sub(r"[^0-9a-z\u4e00-\u9fff-]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    if not text:
        text = "unknown"
    return f"{prefix}-{text}"


def normalize_title_for_dedupe(value):
    text = clean_text(value).lower()
    text = text.replace("：", ":")
    return re.sub(r"[^0-9a-z]+", "", text)


def get_bib_field(bibtex, field):
    if not bibtex:
        return ""
    pattern = re.compile(
        rf"\b{re.escape(field)}\s*=\s*(\{{(?P<braced>.*?)\}}|\"(?P<quoted>.*?)\")\s*,",
        re.I | re.S,
    )
    match = pattern.search(bibtex)
    if not match:
        return ""
    value = match.group("braced") if match.group("braced") is not None else match.group("quoted")
    return clean_text(value)


def bib_key(bibtex):
    match = re.search(r"@\w+\s*\{\s*([^,\s]+)", bibtex or "")
    return clean_text(match.group(1)) if match else ""


def infer_url(bibtex, explicit_url=""):
    explicit_url = clean_text(explicit_url)
    if explicit_url:
        return explicit_url
    url = get_bib_field(bibtex, "url")
    if url:
        return url
    doi = get_bib_field(bibtex, "doi")
    if not doi:
        match = DOI_RE.search(bibtex or "")
        doi = match.group(0) if match else ""
    doi = doi.rstrip("}. ,")
    return f"https://doi.org/{doi}" if doi else ""


def infer_special_tags(title, is_colored_title):
    if not is_colored_title:
        return []
    text = title.lower()
    if "RAG" in title or "retrieval-augmented" in text or "retrieval augmented" in text:
        return ["RAG"]
    return ["Agent"]


def is_marked_title(cell):
    color = cell.font.color
    if color is None:
        return False
    if color.type == "rgb":
        return color.rgb not in (None, "FF000000", "00000000")
    return False


def image_map(ws):
    result = {}
    for image in getattr(ws, "_images", []):
        anchor = image.anchor
        try:
            row = anchor._from.row + 1
            col = anchor._from.col + 1
        except AttributeError:
            continue
        if col != COL["screenshot"] or row in result:
            continue
        result[row] = image
    return result


def save_thumbnail(image, entry_id):
    raw = image._data()
    with Image.open(BytesIO(raw)) as source:
        source = source.convert("RGB")
        for size, directory in ((100, THUMBS100_DIR), (200, THUMBS200_DIR)):
            thumb = Image.new("RGB", (size, size), "white")
            copy = source.copy()
            copy.thumbnail((size, size), Image.Resampling.LANCZOS)
            x = (size - copy.width) // 2
            y = (size - copy.height) // 2
            thumb.paste(copy, (x, y))
            thumb.save(directory / f"{entry_id}.png")


def save_placeholder(entry_id, text):
    for size, directory in ((100, THUMBS100_DIR), (200, THUMBS200_DIR)):
        image = Image.new("RGB", (size, size), "#f4f6f8")
        draw = ImageDraw.Draw(image)
        draw.rectangle((0, 0, size - 1, size - 1), outline="#c7cdd4")
        label = "Vis4VL"
        font = ImageFont.load_default()
        bbox = draw.textbbox((0, 0), label, font=font)
        draw.text(((size - (bbox[2] - bbox[0])) / 2, size / 2 - 8), label, fill="#69717a", font=font)
        image.save(directory / f"{entry_id}.png")


def safe_write_json(path, data):
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def make_reference(entry):
    authors = entry.get("authors") or "Unknown authors"
    title = entry["title"]
    venue = entry.get("venue") or "Unknown venue"
    year = entry.get("year") or ""
    return f"{authors}. <i>{title}</i>. {venue}, {year}."


def clear_generated_outputs():
    DATA_DIR.mkdir(exist_ok=True)
    BIBTEX_DIR.mkdir(exist_ok=True)
    THUMBS100_DIR.mkdir(exist_ok=True)
    THUMBS200_DIR.mkdir(exist_ok=True)
    for directory in (BIBTEX_DIR, THUMBS100_DIR, THUMBS200_DIR):
        for path in directory.glob("*"):
            if path.is_file():
                path.unlink()


def build_categories(rows):
    values_by_group = defaultdict(Counter)
    labels = {}
    for row in rows:
        for key, title, _description in CATEGORY_GROUPS:
            for value in row.get(key, []):
                slug = slugify(value, key)
                values_by_group[key][slug] += 1
                labels[slug] = value

    categories = []
    for key, title, description in CATEGORY_GROUPS:
        entries = []
        order = {value: index for index, value in enumerate(CATEGORY_ORDER.get(key, []))}
        sorted_items = sorted(
            values_by_group[key].items(),
            key=lambda item: (order.get(labels[item[0]], len(order)), labels[item[0]].lower()),
        )
        for slug, _count in sorted_items:
            label = labels[slug]
            entries.append({
                "type": "category-entry",
                "title": slug,
                "description": label,
                "content": label,
            })
        if entries:
            categories.append({
                "type": "category",
                "title": key,
                "description": description,
                "childrenDescription": f"{description}: ",
                "entries": entries,
            })
    return categories


def main():
    xlsx_files = sorted(path for path in XLSX_DIR.glob("*.xlsx") if not path.name.startswith("~$"))
    if not xlsx_files:
        raise FileNotFoundError("No .xlsx file found in vis_data")

    workbook = openpyxl.load_workbook(xlsx_files[0], read_only=False, data_only=True)
    worksheet = workbook.active
    images = image_map(worksheet)
    clear_generated_outputs()

    rows = []
    id_counts = Counter()
    missing_images = []
    marked_counts = Counter()
    seen_titles = {}
    skipped_duplicates = []

    for row_number in range(3, worksheet.max_row + 1):
        title = clean_text(worksheet.cell(row_number, COL["title"]).value)
        if not title:
            continue
        normalized_title = normalize_title_for_dedupe(title)
        if normalized_title in seen_titles:
            skipped_duplicates.append((row_number, seen_titles[normalized_title], title))
            continue
        seen_titles[normalized_title] = row_number

        bibtex = clean_text(worksheet.cell(row_number, COL["bibtex"]).value)
        key = bib_key(bibtex)
        if not key:
            key = re.sub(r"[^A-Za-z0-9]+", "", title.title())[:32] or f"paper{row_number}"
        id_counts[key] += 1
        entry_id = key if id_counts[key] == 1 else f"{key}-{id_counts[key]}"

        marked = is_marked_title(worksheet.cell(row_number, COL["title"]))
        special_tags = infer_special_tags(title, marked)
        for tag in special_tags:
            marked_counts[tag] += 1

        row = {
            "id": entry_id,
            "title": title,
            "url": clean_text(worksheet.cell(row_number, COL["url"]).value),
            "venue": clean_text(worksheet.cell(row_number, COL["venue"]).value),
            "year": int(worksheet.cell(row_number, COL["year"]).value),
            "authors": get_bib_field(bibtex, "author"),
            "keywords": clean_text(worksheet.cell(row_number, COL["keywords"]).value),
            "Keywords": clean_text(worksheet.cell(row_number, COL["keywords"]).value),
            "abstract": clean_text(worksheet.cell(row_number, COL["abstract"]).value),
            "rank": clean_text(worksheet.cell(row_number, COL["rank"]).value),
            "citations": clean_text(worksheet.cell(row_number, COL["citations"]).value),
            "modal_count": clean_text(worksheet.cell(row_number, COL["modal_count"]).value),
            "modality": translated_values(worksheet.cell(row_number, COL["modality"]).value),
            "dataset": clean_text(worksheet.cell(row_number, COL["dataset"]).value),
            "domain": translated_values(worksheet.cell(row_number, COL["domain"]).value),
            "salon": translated_values(worksheet.cell(row_number, COL["salon"]).value),
            "task": translated_values(worksheet.cell(row_number, COL["task"]).value),
            "model": translated_values(worksheet.cell(row_number, COL["model"]).value),
            "anchor_summary": clean_text(worksheet.cell(row_number, COL["anchor_summary"]).value),
            "improved_model": clean_text(worksheet.cell(row_number, COL["improved_model"]).value),
            "fine_tuning": clean_text(worksheet.cell(row_number, COL["fine_tuning"]).value),
            "vis_encoding": translated_values(worksheet.cell(row_number, COL["vis_encoding"]).value),
            "specific_view": clean_text(worksheet.cell(row_number, COL["specific_view"]).value),
            "interaction": translated_values(worksheet.cell(row_number, COL["interaction"]).value),
            "interaction_detail": clean_text(worksheet.cell(row_number, COL["interaction_detail"]).value),
            "evaluation": translated_values(worksheet.cell(row_number, COL["evaluation"]).value),
            "special_tags": special_tags,
            "is_marked": marked,
        }

        category_values = []
        for key_name, _title, _description in CATEGORY_GROUPS:
            category_values.extend(slugify(value, key_name) for value in row.get(key_name, []))
        row["categories"] = category_values
        row["reference"] = make_reference(row)

        if bibtex:
            (BIBTEX_DIR / f"{entry_id}.bib").write_text(bibtex + "\n", encoding="utf-8")
        else:
            (BIBTEX_DIR / f"{entry_id}.bib").write_text("", encoding="utf-8")

        if row_number in images:
            save_thumbnail(images[row_number], entry_id)
        else:
            save_placeholder(entry_id, title)
            missing_images.append(row_number)

        rows.append(row)

    categories = build_categories(rows)
    safe_write_json(DATA_DIR / "content.json", rows)
    safe_write_json(DATA_DIR / "categories.json", categories)

    print(f"Generated {len(rows)} entries")
    print(f"Generated {sum(len(group['entries']) for group in categories)} category entries in {len(categories)} groups")
    print(f"Special tags: {dict(marked_counts)}")
    print(f"Missing images: {missing_images}")
    print(f"Skipped duplicate titles: {skipped_duplicates}")


if __name__ == "__main__":
    main()
