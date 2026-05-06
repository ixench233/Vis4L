import json
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from difflib import SequenceMatcher
from pathlib import Path
from urllib.parse import quote

import requests
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = next((ROOT / "vis_data").glob("*.xlsx"))
CACHE_PATH = ROOT / "data" / "paper_pdf_url_cache.json"
OLD_URL_CACHE_PATH = ROOT / "data" / "paper_url_cache.json"
MIN_SCORE = 0.78


def norm_title(value):
    text = str(value or "").lower()
    text = text.replace("：", ":").replace("–", "-").replace("—", "-").replace("‐", "-")
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def title_similarity(a, b):
    a = norm_title(a)
    b = norm_title(b)
    if not a or not b:
        return 0
    if a == b:
        return 1
    return SequenceMatcher(None, a, b).ratio()


def year_bonus(expected_year, candidate_year):
    try:
        if expected_year and candidate_year and abs(int(expected_year) - int(candidate_year)) <= 1:
            return 0.04
    except Exception:
        pass
    return 0


def is_pdf_url(url):
    url = str(url or "").strip()
    return bool(url) and (".pdf" in url.lower() or "pdf" in url.lower())


def normalize_pdf_url(url):
    url = str(url or "").strip()
    if url.startswith("http://doi.org/"):
        return "https://doi.org/" + url.removeprefix("http://doi.org/")
    return url


def load_json(path):
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return {}


def doi_from_old_cache(key):
    old_cache = load_json(OLD_URL_CACHE_PATH)
    url = old_cache.get(key, {}).get("url", "")
    match = re.search(r"10\.\d{4,9}/\S+", url, re.I)
    if not match:
        return ""
    return match.group(0).rstrip(").,")


def candidate_result(url, score, matched_title, source):
    return {
        "url": normalize_pdf_url(url),
        "score": round(score, 4),
        "matched_title": matched_title or "",
        "source": source,
    }


def best_openalex_pdf(title, year):
    response = requests.get(
        "https://api.openalex.org/works",
        params={"search": title, "per-page": 5},
        timeout=12,
        headers={"User-Agent": "vis4l-pdf-url-fill/1.0"},
    )
    response.raise_for_status()

    best = None
    for work in response.json().get("results", []):
        matched_title = work.get("title") or ""
        score = title_similarity(title, matched_title) + year_bonus(year, work.get("publication_year"))
        urls = []
        for key in ("best_oa_location", "primary_location"):
            location = work.get(key) or {}
            urls.append(location.get("pdf_url"))
            urls.append(location.get("landing_page_url"))
        for location in work.get("locations") or []:
            urls.append((location or {}).get("pdf_url"))
            urls.append((location or {}).get("landing_page_url"))

        for url in urls:
            if score >= MIN_SCORE and is_pdf_url(url):
                result = candidate_result(url, score, matched_title, "openalex")
                if not best or result["score"] > best["score"]:
                    best = result

    return best


def best_semantic_scholar_pdf(title, year):
    response = requests.get(
        "https://api.semanticscholar.org/graph/v1/paper/search",
        params={
            "query": title,
            "limit": 5,
            "fields": "title,year,openAccessPdf,externalIds,url",
        },
        timeout=12,
        headers={"User-Agent": "vis4l-pdf-url-fill/1.0"},
    )
    response.raise_for_status()

    best = None
    for paper in response.json().get("data") or []:
        matched_title = paper.get("title") or ""
        score = title_similarity(title, matched_title) + year_bonus(year, paper.get("year"))
        pdf = paper.get("openAccessPdf") or {}
        url = pdf.get("url")
        if score >= MIN_SCORE and is_pdf_url(url):
            result = candidate_result(url, score, matched_title, "semantic_scholar")
            if not best or result["score"] > best["score"]:
                best = result

    return best


def crossref_pdf_from_doi(doi, title, year):
    if not doi:
        return None
    response = requests.get(
        "https://api.crossref.org/works/" + quote(doi, safe=""),
        timeout=12,
        headers={"User-Agent": "vis4l-pdf-url-fill/1.0"},
    )
    response.raise_for_status()

    message = response.json().get("message", {})
    matched_title = (message.get("title") or [""])[0]
    score = title_similarity(title, matched_title)
    issued = message.get("issued", {}).get("date-parts", [[]])[0]
    if issued:
        score += year_bonus(year, issued[0])
    for link in message.get("link") or []:
        url = link.get("URL")
        content_type = str(link.get("content-type") or "").lower()
        if score >= MIN_SCORE and (content_type == "application/pdf" or is_pdf_url(url)):
            return candidate_result(url, score, matched_title, "crossref")
    return None


def lookup_pdf(item):
    key, title, year = item
    errors = []

    try:
        result = best_openalex_pdf(title, year)
        if result:
            return key, result
    except Exception as exc:
        errors.append("openalex: " + str(exc))

    try:
        result = best_semantic_scholar_pdf(title, year)
        if result:
            return key, result
    except Exception as exc:
        errors.append("semantic_scholar: " + str(exc))

    try:
        result = crossref_pdf_from_doi(doi_from_old_cache(key), title, year)
        if result:
            return key, result
    except Exception as exc:
        errors.append("crossref: " + str(exc))

    return key, {"url": "", "score": 0, "matched_title": "", "source": "", "errors": errors[:3]}


def collect_titles(worksheet):
    seen = {}
    items = []
    for row in range(3, worksheet.max_row + 1):
        title = worksheet.cell(row, 2).value
        if not title:
            continue
        key = norm_title(title)
        if key in seen:
            continue
        seen[key] = True
        items.append((key, str(title), worksheet.cell(row, 4).value))
    return items


def find_or_create_url_column(worksheet):
    for col in range(1, worksheet.max_column + 1):
        if str(worksheet.cell(2, col).value or "").strip().lower() == "url":
            return col
    return worksheet.max_column + 1


def main():
    cache = load_json(CACHE_PATH)
    workbook = load_workbook(XLSX_PATH)
    worksheet = workbook.active
    items = collect_titles(worksheet)
    pending = [item for item in items if item[0] not in cache]

    print(f"titles: {len(items)}, cached: {len(cache)}, pending: {len(pending)}")
    completed = 0
    with ThreadPoolExecutor(max_workers=6) as executor:
        futures = [executor.submit(lookup_pdf, item) for item in pending]
        for future in as_completed(futures):
            key, result = future.result()
            cache[key] = result
            completed += 1
            if completed % 25 == 0:
                CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
                print(f"queried {completed}/{len(pending)}")
            time.sleep(0.02)

    CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")

    url_col = find_or_create_url_column(worksheet)
    worksheet.cell(1, url_col).value = "论文地址"
    worksheet.cell(2, url_col).value = "URL"
    worksheet.column_dimensions[worksheet.cell(1, url_col).column_letter].width = 44

    filled = 0
    blank = 0
    for row in range(3, worksheet.max_row + 1):
        title = worksheet.cell(row, 2).value
        if not title:
            continue
        url = cache.get(norm_title(title), {}).get("url", "")
        cell = worksheet.cell(row, url_col)
        if url:
            cell.value = url
            cell.hyperlink = url
            cell.style = "Hyperlink"
            filled += 1
        else:
            cell.value = None
            blank += 1

    workbook.save(XLSX_PATH)
    print(f"saved: {XLSX_PATH}")
    print(f"filled_pdf_urls: {filled}, blank: {blank}, column: {worksheet.cell(1, url_col).column_letter}")


if __name__ == "__main__":
    main()
